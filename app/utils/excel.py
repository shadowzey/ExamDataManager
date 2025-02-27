"""
Excel 处理工具模块
"""
import io
import pandas as pd
import math
from typing import List, Dict, Any, Tuple
from openpyxl.styles import Alignment, PatternFill
import asyncio
from concurrent.futures import ThreadPoolExecutor


# 列名常量
COLUMNS = ["编号", "姓名", "电话", "收款账号身份证号", "收款账号", "开户行", "收款账号开户行", 
          "电话号码", "备注", "身份证号", "银行卡号", "发放明细", "次数（小时）", "标准", 
          "其他项目", "工资编号", "分院", "金额"]

# 列名映射
COLUMNS_MAPPING = {
    "姓名": "name",
    "收款账号身份证号": "id_card",
    "收款账号": "bank_card",
    "收款账号开户行": "bank_name",
    "电话号码": "phone",
    "备注": "remark",
    "身份证号": "id_card",
    "银行卡号": "bank_card",
    "发放明细": "detail",
    "次数（小时）": "count_hour",
    "标准": "standard",
    "其他项目": "other_project",
    "电话": "phone",
    "开户行": "bank_name",
    "工资编号": "salary_id",
    "分院": "department"
}

# 单元格颜色常量
REPEAT_COLOR = "FFFF0000"  # 红色
NO_MATCH_COLOR = "FFFFFF00"  # 黄色

# 创建线程池执行器
_thread_pool = ThreadPoolExecutor()


def read_excel_to_dict_list(contents: bytes, sheet_name: str, header: int = 0) -> List[Dict[str, Any]]:
    """
    读取 Excel 文件内容并转换为字典列表
    
    Args:
        contents: Excel 文件的二进制内容
        sheet_name: 要读取的工作表名称
        header: 表头行索引，默认为0
        
    Returns:
        包含 Excel 数据的字典列表
    """
    # 读取 Excel 文件中的指定 sheet，并将第一行作为列标题
    df = pd.read_excel(io.BytesIO(contents), sheet_name=sheet_name, header=header)
    df = handle_data(df)
    # 将数据框转换为字典的列表
    data_dict_list = df.to_dict(orient='records')

    return data_dict_list


def handle_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    处理 DataFrame 数据
    
    Args:
        df: 输入的 DataFrame
        
    Returns:
        处理后的 DataFrame
    """
    # 检查DataFrame是否为空
    if df.empty:
        return df
    
    # 检查是否有'name'列，如果没有但有'姓名'列，则使用'姓名'列
    name_column = None
    if 'name' in df.columns:
        name_column = 'name'
    elif '姓名' in df.columns:
        name_column = '姓名'
    
    # 如果找到了姓名列，过滤掉姓名为空的行
    if name_column:
        df = df[df[name_column].notna() & (df[name_column] != '')]
    
    # 获取列顺序
    order_index = {value: index for index, value in enumerate(list(df.columns))}
    
    # 检查是否有需要处理的列
    if not set(COLUMNS).intersection(df.columns) and not set(COLUMNS_MAPPING.keys()).intersection(df.columns):
        # 如果没有匹配的列，直接返回原始DataFrame
        return df
    
    # 和 COLUMNS 取交集
    columns = set(COLUMNS).intersection(df.columns)
    
    # 如果没有匹配的列，尝试使用英文列名
    if not columns:
        # 直接返回所有列
        return df
    
    # 按照预定义的顺序排序列
    sorted_list = sorted(list(columns), key=lambda x: order_index.get(x, float('inf')))
    
    # 确保排序列表不为空
    if sorted_list:
        df = df[sorted_list]
    
    return df


async def write_dict_list_to_excel(
    data_dict_list: List[Dict[str, Any]], 
    sheet_name: str, 
    repeat_list: List[List[int]], 
    no_match_list: List[int]
) -> bytes:
    """
    将字典列表写入 Excel 文件（异步版本）
    
    Args:
        data_dict_list: 要写入的数据字典列表
        sheet_name: 工作表名称
        repeat_list: 重复项索引列表
        no_match_list: 未匹配项索引列表
        
    Returns:
        Excel 文件的二进制内容
    """
    # 使用线程池执行IO密集型操作
    return await asyncio.get_event_loop().run_in_executor(
        _thread_pool, 
        _write_excel_in_thread,
        data_dict_list, 
        sheet_name, 
        repeat_list, 
        no_match_list
    )


def _write_excel_in_thread(
    data_dict_list: List[Dict[str, Any]], 
    sheet_name: str, 
    repeat_list: List[List[int]], 
    no_match_list: List[int]
) -> bytes:
    """
    在线程中执行Excel写入操作
    
    Args:
        data_dict_list: 要写入的数据字典列表
        sheet_name: 工作表名称
        repeat_list: 重复项索引列表
        no_match_list: 未匹配项索引列表
        
    Returns:
        Excel 文件的二进制内容
    """
    # 将字典列表转换为 DataFrame
    df = pd.DataFrame(data_dict_list)
    output = io.BytesIO()
    min_length = 14
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name, float_format="%.2f")
        # 获取 openpyxl workbook 对象
        workbook = writer.book
        worksheet = workbook[sheet_name]

        # 获取列索引并转换为 Excel 列字母
        sp_col_idx = df.columns.get_loc('发放明细') + 1 if '发放明细' in df.columns else None
        # 获取金额列
        je_col_idx = df.columns.get_loc('金额') + 1 if '金额' in df.columns else None

        # 调整列宽
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) + min_length for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length
            
            for cell in worksheet[1]:  # 第一行通常是表头
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置重复项的背景色
        red_fill = PatternFill(start_color=REPEAT_COLOR, end_color=REPEAT_COLOR, fill_type="solid")
        for repeat_indices in repeat_list:
            for idx in repeat_indices:
                try:
                    # 确保idx是整数
                    idx_num = int(idx) if isinstance(idx, str) else idx
                    # 索引需要+2，因为Excel从1开始，且有表头行
                    row_idx = idx_num + 2  # 修改为+2，因为有表头行
                    for col_idx in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = red_fill
                except (ValueError, TypeError) as e:
                    print(f"处理重复项时出错: {e}, idx类型: {type(idx)}, 值: {idx}")
        
        # 设置未匹配项的背景色
        yellow_fill = PatternFill(start_color=NO_MATCH_COLOR, end_color=NO_MATCH_COLOR, fill_type="solid")
        for idx in no_match_list:
            try:
                # 确保idx是整数
                idx_num = int(idx) if isinstance(idx, str) else idx
                # 索引需要+2，因为Excel从1开始，且有表头行
                row_idx = idx_num + 2  # 修改为+2，因为有表头行
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = yellow_fill
            except (ValueError, TypeError) as e:
                print(f"处理未匹配项时出错: {e}, idx类型: {type(idx)}, 值: {idx}")
        
        # 保存工作簿
        workbook.save(output)
    
    # 将文件指针移到开始位置
    output.seek(0)
    return output.getvalue()


async def read_excel_to_dict_list_async(contents: bytes, sheet_name: str, header: int = 0) -> List[Dict[str, Any]]:
    """
    异步读取 Excel 文件内容并转换为字典列表
    
    Args:
        contents: Excel 文件的二进制内容
        sheet_name: 要读取的工作表名称
        header: 表头行索引，默认为0
        
    Returns:
        包含 Excel 数据的字典列表
    """
    # 使用线程池执行IO密集型操作
    return await asyncio.get_event_loop().run_in_executor(
        _thread_pool, 
        read_excel_to_dict_list,
        contents, 
        sheet_name, 
        header
    )
